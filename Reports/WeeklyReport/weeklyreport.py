import pandas as pd
import mysql.connector
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from typing import Dict, List
import re
import paramiko
import logging
import os
import sys
from datetime import datetime
import traceback
import time
import openpyxl.styles
import subprocess
import concurrent.futures
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from pathlib import Path

# Set up logging with timestamp in filename
current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
log_filename = f'weekly_report_{current_time}.log'

# Create logs directory if it doesn't exist
os.makedirs('logs', exist_ok=True)
log_filepath = os.path.join('logs', log_filename)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filepath),
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger(__name__)

def setup_tunnel():
    """Set up SSH tunnel using system SSH command"""
    ssh_host = os.environ['SSH_HOST']
    ssh_user = os.environ['SSH_USERNAME']
    mysql_host = os.environ['MYSQL_HOST']
    
    # Build SSH command
    cmd = f"ssh -v -N -L 3307:{mysql_host}:3306 {ssh_user}@{ssh_host}"
    
    # Log without sensitive information
    logger.info("Starting SSH tunnel...")
    
    # Start tunnel in background
    process = subprocess.Popen(
        cmd.split(),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE
    )
    
    # Wait a bit for tunnel to establish
    time.sleep(5)
    return process

def send_email(file_path: str):
    """Send email with the report as attachment"""
    try:
        logger.info("Preparing to send email...")
        
        # Email settings
        sender_email = os.environ['GMAIL_USER']
        sender_password = os.environ['GMAIL_APP_PASSWORD']
        
        # Get recipients from environment variable
        try:
            recipients_str = os.environ['WEEKLY_REPORT_RECEPIENTS']
            # Clean up the string and extract emails
            recipients = [email.strip().strip("'[]") for email in recipients_str.replace(' ', '').split(',')]
            recipients = [email for email in recipients if '@' in email]  # Validate emails
            logger.info(f"Found {len(recipients)} recipients")
            if not recipients:
                raise ValueError("No valid email addresses found in WEEKLY_REPORT_RECEPIENTS")
        except KeyError:
            logger.error("WEEKLY_REPORT_RECEPIENTS environment variable not set")
            raise
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = 'Weekly Report'
        
        # Add body
        body = """Hi Team,

Please find the attached report.

Regards,
AK"""
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach file
        with open(file_path, 'rb') as f:
            attachment = MIMEApplication(f.read(), _subtype='xlsx')
            attachment.add_header('Content-Disposition', 'attachment', 
                                filename=os.path.basename(file_path))
            msg.attach(attachment)
        
        # Send email
        logger.info("Connecting to Gmail SMTP server...")
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
            
        logger.info("Email sent successfully")
        
    except Exception as e:
        # Ensure error messages don't contain sensitive info
        error_msg = str(e)
        if any(secret in error_msg for secret in [sender_email, sender_password] + recipients):
            error_msg = "Email error occurred (sensitive information redacted)"
        logger.error(f"Error sending email: {error_msg}")
        logger.error("Stack trace omitted for security")
        raise

class ReportGenerator:
    def __init__(self, db_config: Dict):
        self.db_config = db_config
        self.regions = {
            'Dubai': ['adarsh', 'sagun', 'simrahashraf', 'raghav', 'nihad', 'waseem'],
            'Abu Dhabi': ['clarita', 'adhil', 'sana', 'lamis'],
            'Sharjah': ['taiba', 'sahbhan', 'akef', 'aaftab'],
            'KAM': ['musir', 'sparsh', 'sourabh', 'tariq']
        }
        self.tunnel_process = None
        self._init_tunnel()

    def _init_tunnel(self):
        """Initialize SSH tunnel with retries"""
        max_retries = 3
        retry_count = 0
        while retry_count < max_retries:
            try:
                logger.info(f"Attempting to create SSH tunnel (attempt {retry_count + 1}/{max_retries})")
                self.tunnel_process = setup_tunnel()
                if self.tunnel_process.poll() is None:
                    logger.info("SSH tunnel successfully created and active")
                    break
                logger.warning("SSH tunnel process failed")
            except Exception as e:
                retry_count += 1
                if retry_count == max_retries:
                    logger.error("Failed to create SSH tunnel after maximum retries")
                    raise
                logger.warning(f"SSH tunnel creation failed, retrying... Error: {str(e)}")
                time.sleep(2)

    def get_db_connection(self) -> mysql.connector.connection.MySQLConnection:
        """Get database connection through SSH tunnel"""
        try:
            if self.tunnel_process is None or self.tunnel_process.poll() is not None:
                logger.error("SSH tunnel is not active")
                self._init_tunnel()
            
            logger.info("Attempting to establish database connection...")
            
            # Use environment variables for connection
            conn = mysql.connector.connect(
                host='127.0.0.1',
                port=3307,
                user=os.environ['MYSQL_USER'],
                password=os.environ['MYSQL_PASSWORD'],
                database=os.environ['MYSQL_DATABASE'],
                connection_timeout=10
            )
            
            # Test the connection
            cursor = conn.cursor(buffered=True)
            cursor.execute("SELECT 1")
            result = cursor.fetchone()
            cursor.close()
            
            if result != (1,):
                raise Exception("Connection test failed")
                
            logger.info("Database connection established and tested successfully")
            return conn
            
        except mysql.connector.Error as e:
            # Redact sensitive information from error messages
            error_msg = str(e)
            if any(secret in error_msg for secret in [os.environ.get('MYSQL_USER', ''), 
                                                    os.environ.get('MYSQL_PASSWORD', ''),
                                                    os.environ.get('MYSQL_DATABASE', '')]):
                error_msg = "Database error occurred (sensitive information redacted)"
            logger.error(f"MySQL connection error: {error_msg}")
            logger.error("Error details omitted for security")
            raise
        except Exception as e:
            logger.error("Unexpected error in database connection (details omitted for security)")
            raise

    def __del__(self):
        """Cleanup method to ensure tunnel process is terminated"""
        if self.tunnel_process:
            try:
                self.tunnel_process.terminate()
                self.tunnel_process.wait(timeout=5)
                logger.info("SSH tunnel terminated")
            except Exception as e:
                logger.error(f"Error terminating SSH tunnel: {str(e)}")
                try:
                    self.tunnel_process.kill()
                    logger.info("SSH tunnel force killed")
                except:
                    pass

    def validate_excel_template(self, template_path: str) -> None:
        """Validate Excel template existence and structure"""
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Excel template not found: {template_path}")
        
        try:
            workbook = openpyxl.load_workbook(template_path)
            logger.info(f"Successfully opened Excel template: {template_path}")
            
            # Only need to verify the main sheet exists
            if 'Sheet1' not in workbook.sheetnames:
                raise ValueError("Template must contain 'Sheet1'")
            
            logger.info("Excel template validated successfully")
            workbook.close()
        except Exception as e:
            logger.error(f"Error validating Excel template: {str(e)}")
            raise

    def modify_query_for_region(self, query: str, region: str) -> str:
        try:
            agents = self.regions[region]
            agents_str = "','".join(agents)
            logger.info(f"Modifying query for region {region} with agents: {agents_str}")
            
            pattern = r"ram\.agent_name\s+in\s*\([^)]+\)"
            new_filter = f"ram.agent_name in ('{agents_str}')"
            
            # Log the original query
            logger.info("Original query:")
            logger.info(query)
            
            modified_query = re.sub(pattern, new_filter, query)
            
            # Log the modified query
            logger.info("Modified query:")
            logger.info(modified_query)
            
            # Verify the modification was successful
            if modified_query == query:
                logger.warning("Query modification might have failed - no changes detected")
            
            return modified_query
        except Exception as e:
            logger.error(f"Error modifying query for {region}: {str(e)}")
            logger.error(traceback.format_exc())
            raise

    def execute_query(self, query: str, region: str) -> pd.DataFrame:
        conn = None
        attempts = 0
        max_attempts = 3
        
        # Extract only the SQL part of the query
        sql_query = self._extract_sql_from_query(query)
        
        while attempts < max_attempts:
            try:
                logger.info(f"Processing region: {region}")
                conn = self.get_db_connection()
                modified_query = self.modify_query_for_region(sql_query, region)
                logger.info(f"Executing modified query for {region}...")
                
                # Print the actual modified query for debugging
                logger.info("Modified Query:")
                logger.info(modified_query)
                
                # Execute with cursor first to test query
                cursor = conn.cursor(buffered=True)
                logger.info("Executing query with cursor first...")
                cursor.execute(modified_query)
                logger.info("Query executed successfully with cursor")
                
                # Fetch results with pandas
                df = pd.read_sql_query(modified_query, conn)
                
                logger.info(f"Query execution successful for {region} - Retrieved {len(df)} rows")
                if df.empty:
                    logger.warning(f"Query returned empty DataFrame for {region}")
                else:
                    logger.info(f"DataFrame columns: {df.columns.tolist()}")
                    logger.info(f"First row sample: {df.iloc[0].to_dict()}")
                return df
                
            except mysql.connector.Error as e:
                attempts += 1
                logger.error(f"MySQL Error for {region} (attempt {attempts}/{max_attempts}): {str(e)}")
                logger.error(f"Error details: {e.__class__.__name__}")
                if attempts == max_attempts:
                    raise
                time.sleep(5)
            except Exception as e:
                logger.error(f"Unexpected error executing query for {region}: {str(e)}")
                logger.error(f"Error type: {type(e).__name__}")
                logger.error(traceback.format_exc())
                raise
            finally:
                if conn:
                    try:
                        conn.close()
                        logger.debug(f"Database connection closed for {region}")
                    except Exception as e:
                        logger.warning(f"Error closing database connection for {region}: {str(e)}")

    def _extract_sql_from_query(self, query: str) -> str:
        """Extract the SQL part from the query text by finding the first SELECT statement."""
        try:
            # Find the position of the first SELECT statement
            select_pos = query.upper().find('SELECT')
            if select_pos == -1:
                raise ValueError("No SELECT statement found in query")
            
            # Return everything from SELECT onwards
            sql_query = query[select_pos:]
            logger.info("Successfully extracted SQL query")
            return sql_query
        except Exception as e:
            logger.error(f"Error extracting SQL from query: {str(e)}")
            logger.error(f"Original query text: {query[:200]}...")  # Log first 200 chars
            raise

    def process_region(self, region: str, queries: Dict[str, str]) -> tuple:
        results = {}
        try:
            logger.info(f"Starting to process region: {region}")
            logger.info(f"Number of queries to process: {len(queries)}")
            
            def execute_single_query(query_tuple):
                query_name, query = query_tuple
                max_retries = 3
                retry_count = 0
                
                while retry_count < max_retries:
                    try:
                        logger.info(f"Processing query '{query_name}' for region '{region}'...")
                        result = self.execute_query(query, region)
                        logger.info(f"Successfully completed query '{query_name}' for region '{region}'")
                        return query_name, result
                    except Exception as e:
                        retry_count += 1
                        logger.error(f"Attempt {retry_count} failed for query '{query_name}' in region '{region}': {str(e)}")
                        if retry_count == max_retries:
                            raise
                        time.sleep(5)  # Wait before retry
                
            # Execute queries in parallel using ThreadPoolExecutor
            with ThreadPoolExecutor(max_workers=7) as executor:
                # Submit all queries for execution
                future_to_query = {
                    executor.submit(execute_single_query, (query_name, query)): query_name 
                    for query_name, query in queries.items()
                }
                
                # Collect results as they complete
                for future in concurrent.futures.as_completed(future_to_query):
                    query_name = future_to_query[future]
                    try:
                        query_name, result = future.result()
                        results[query_name] = result
                    except Exception as e:
                        logger.error(f"Query '{query_name}' failed for region '{region}': {str(e)}")
                        raise
            
            logger.info(f"Completed processing all queries for region: {region}")
            return region, results
        except Exception as e:
            logger.error(f"Error processing region {region}: {str(e)}")
            logger.error(f"Error type: {type(e).__name__}")
            logger.error(traceback.format_exc())
            raise

    def update_excel_template(self, template_path: str, results: Dict[str, Dict[str, pd.DataFrame]], output_file: str):
        try:
            logger.info(f"Opening Excel template: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            worksheet = workbook['Sheet1']  # Use the single sheet
            
            # Process results for each region
            for region, region_results in results.items():
                logger.info(f"\nProcessing region: {region}")
                self._process_query_results(worksheet, region_results, region)

            try:
                workbook.save(output_file)
                logger.info(f"Excel file successfully saved to {output_file}")
            except Exception as e:
                logger.error(f"Error saving workbook: {str(e)}")
                raise
            
        except Exception as e:
            logger.error(f"Error updating Excel template: {str(e)}")
            logger.error(traceback.format_exc())
            raise

    def _process_query_results(self, worksheet, region_results, region: str):
        """Process query results and update worksheet"""
        try:
            logger.info(f"Processing results for region: {region}")
            
            # Define column mappings for each region
            column_mappings = {
                'Dubai': {'data': 'C', 'prev': 'D', 'growth': 'E'},
                'Abu Dhabi': {'data': 'F', 'prev': 'G', 'growth': 'H'},
                'Sharjah': {'data': 'I', 'prev': 'J', 'growth': 'K'},
                'KAM': {'data': 'L', 'prev': 'M', 'growth': 'N'}
            }
            
            cols = column_mappings[region]

            # Update header rows
            def update_headers():
                # Region headers (Row 1)
                region_ranges = {
                    'Dubai': 'C1:E1',
                    'Abu Dhabi': 'F1:H1',
                    'Sharjah': 'I1:K1',
                    'KAM': 'L1:N1'
                }
                
                # Unmerge all cells in header rows first
                for row in [1, 2]:
                    for merge_range in worksheet.merged_cells.ranges.copy():
                        if merge_range.min_row == row:
                            worksheet.unmerge_cells(str(merge_range))

                # Set region headers and merge cells
                for reg, cell_range in region_ranges.items():
                    start_col, end_col = cell_range.split(':')[0][0], cell_range.split(':')[1][0]
                    worksheet.merge_cells(cell_range)
                    worksheet[f'{start_col}1'] = reg
                    
                    # Week headers (Row 2)
                    worksheet[f'{start_col}2'] = 'Week'
                    worksheet[f'{chr(ord(start_col) + 1)}2'] = 'Week'
                    worksheet[f'{end_col}2'] = 'Growth/Degrowth %'
                    
                    # Previous headers (Row 3)
                    worksheet[f'{start_col}3'] = '(Previous)'
                    worksheet[f'{chr(ord(start_col) + 1)}3'] = '(Previous to Previous)'
                    
                    # Merge Week cells vertically in row 2-3
                    worksheet.merge_cells(f'{start_col}2:{start_col}3')
                    worksheet.merge_cells(f'{chr(ord(start_col) + 1)}2:{chr(ord(start_col) + 1)}3')
                    worksheet.merge_cells(f'{end_col}2:{end_col}3')

                # Apply formatting to headers
                for row in [1, 2, 3]:
                    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                        cell = worksheet[f'{col}{row}']
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center', 
                                                                 vertical='center',
                                                                 wrap_text=True)
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.border = openpyxl.styles.Border(
                            left=openpyxl.styles.Side(style='thin'),
                            right=openpyxl.styles.Side(style='thin'),
                            top=openpyxl.styles.Side(style='thin'),
                            bottom=openpyxl.styles.Side(style='thin')
                        )

            # Update headers only once when processing the first region
            if region == 'Dubai':  # Only do this for the first region
                update_headers()

            # Define row mappings for all parameters
            row_mappings = {
                'Orders': {
                    'Total Orders': 4,  # Updated row numbers to start after headers
                    'Successful Orders': 5,
                    'Rejected Orders': 6,
                    'Rejected Orders %': 7
                },
                'Sales': {
                    'Total Sales': 8,
                    'Net Sales': 9
                },
                'Revenue': {
                    'Commissions': 10,
                    'Payment Gateway': 11
                },
                'Customers': {
                    'Order Frequency': 12,
                    'Smiles Subscription Orders': 13,
                    'New Customer Count': 14,
                    'New Customer Order Count': 15,
                    'Repeat Customer Count': 16,
                    'Repeat Customer Order Count': 17
                },
                'Discounts': {
                    'No Discount Orders': 18,
                    'Restaurant Sponsored Orders': 19,
                    'Smiles Sponsored Orders': 20,
                    'Co-fund orders': 21,
                    'Flat Discount': 22
                }
            }

            def safe_write_cell(ws, cell, value):
                """Safely write to a cell, handling merged cells"""
                try:
                    # Get the master cell if this is part of a merged range
                    cell_coord = cell
                    for merged_range in ws.merged_cells.ranges:
                        if cell in merged_range:
                            cell_coord = merged_range.start_cell.coordinate
                            break
                    
                    # Format numeric values
                    if isinstance(value, (int, float)):
                        if isinstance(value, float) and value % 1 == 0:
                            value = int(value)
                        elif isinstance(value, float):
                            value = round(value, 2)
                    
                    # Write to the cell
                    ws[cell_coord] = value
                    logger.debug(f"Successfully wrote value {value} to cell {cell_coord}")
                    
                except Exception as e:
                    logger.error(f"Error writing to cell {cell}: {str(e)}")
                    raise
            
            # Write parameter names in column B if they don't exist
            for category, params in row_mappings.items():
                for param_name, row in params.items():
                    worksheet[f'B{row}'] = param_name

            for query_name, df in region_results.items():
                logger.info(f"Processing query: {query_name}")
                
                try:
                    # Orders metrics
                    if query_name == 'orders_metrics':
                        for metric, row in row_mappings['Orders'].items():
                            if not df.empty and metric in df['Parameters'].values:
                                row_data = df[df['Parameters'] == metric].iloc[0]
                                safe_write_cell(worksheet, f"{cols['data']}{row}", row_data['Week (Previous)'])
                                safe_write_cell(worksheet, f"{cols['prev']}{row}", row_data['Week (Previous to Previous)'])
                                safe_write_cell(worksheet, f"{cols['growth']}{row}", row_data['Growth/Degrowth %'])

                    # Sales metrics
                    elif query_name == 'sales_metrics':
                        for metric, row in row_mappings['Sales'].items():
                            if not df.empty and metric in df['Parameters'].values:
                                row_data = df[df['Parameters'] == metric].iloc[0]
                                safe_write_cell(worksheet, f"{cols['data']}{row}", row_data['Week (Previous)'])
                                safe_write_cell(worksheet, f"{cols['prev']}{row}", row_data['Week (Previous to Previous)'])
                                safe_write_cell(worksheet, f"{cols['growth']}{row}", row_data['Growth/Degrowth %'])

                    # Commission metrics
                    elif query_name == 'commission_metrics':
                        if not df.empty:
                            commission_row = row_mappings['Revenue']['Commissions']
                            gateway_row = row_mappings['Revenue']['Payment Gateway']
                            
                            safe_write_cell(worksheet, f"{cols['data']}{commission_row}", df.iloc[0]['Previous Week Commission'])
                            safe_write_cell(worksheet, f"{cols['prev']}{commission_row}", df.iloc[0]['Previous to Previous Week Commission'])
                            safe_write_cell(worksheet, f"{cols['growth']}{commission_row}", df.iloc[0]['Commission Growth%'])
                            
                            safe_write_cell(worksheet, f"{cols['data']}{gateway_row}", df.iloc[0]['Previous Week Gateway'])
                            safe_write_cell(worksheet, f"{cols['prev']}{gateway_row}", df.iloc[0]['Previous to Previous Week Gateway'])
                            safe_write_cell(worksheet, f"{cols['growth']}{gateway_row}", df.iloc[0]['Gateway Growth%'])

                    # Order Frequency
                    elif query_name == 'order_frequency':
                        if not df.empty:
                            row = row_mappings['Customers']['Order Frequency']
                            safe_write_cell(worksheet, f"{cols['data']}{row}", df.iloc[0]['Week (Previous)'])
                            safe_write_cell(worksheet, f"{cols['prev']}{row}", df.iloc[0]['Week (Previous to Previous)'])
                            safe_write_cell(worksheet, f"{cols['growth']}{row}", df.iloc[0]['Growth/Degrowth %'])

                    # Subscription Orders
                    elif query_name == 'subscription_orders':
                        if not df.empty:
                            row = row_mappings['Customers']['Smiles Subscription Orders']
                            safe_write_cell(worksheet, f"{cols['data']}{row}", df.iloc[0]['Week (Previous)'])
                            safe_write_cell(worksheet, f"{cols['prev']}{row}", df.iloc[0]['Week (Previous to Previous)'])
                            safe_write_cell(worksheet, f"{cols['growth']}{row}", df.iloc[0]['Growth/Degrowth %'])

                    # New Customers
                    elif query_name == 'new_customers':
                        if not df.empty and len(df) >= 2:
                            try:
                                logger.info(f"Processing new_customers for {region}")
                                logger.info(f"DataFrame columns: {df.columns.tolist()}")
                                logger.info(f"DataFrame data: {df.head().to_dict()}")
                                
                                # Get the values
                                current_value = df.iloc[1]['weekly_new_customers']
                                prev_value = df.iloc[0]['weekly_new_customers']
                                
                                # Calculate growth
                                growth = ((float(current_value) - float(prev_value)) / float(prev_value)) * 100 if prev_value != 0 else 0
                                growth = round(growth, 2)
                                
                                # Update both New Customer Count and New Customer Order Count with the same values
                                for metric in ['New Customer Count', 'New Customer Order Count']:
                                    row = row_mappings['Customers'][metric]
                                    safe_write_cell(worksheet, f"{cols['data']}{row}", current_value)
                                    safe_write_cell(worksheet, f"{cols['prev']}{row}", prev_value)
                                    safe_write_cell(worksheet, f"{cols['growth']}{row}", growth)
                                    
                            except Exception as e:
                                logger.error(f"Error processing new_customers for {region}: {str(e)}")
                                logger.error(traceback.format_exc())

                    # Repeat Customers
                    elif query_name == 'repeat_customers':
                        if not df.empty and len(df) >= 2:
                            try:
                                logger.info(f"Processing repeat_customers for {region}")
                                
                                # Process Repeat Customer Count
                                current_value = df.iloc[1]['weekly_old_customers']
                                prev_value = df.iloc[0]['weekly_old_customers']
                                growth = ((float(current_value) - float(prev_value)) / float(prev_value)) * 100 if prev_value != 0 else 0
                                
                                row = row_mappings['Customers']['Repeat Customer Count']
                                safe_write_cell(worksheet, f"{cols['data']}{row}", current_value)
                                safe_write_cell(worksheet, f"{cols['prev']}{row}", prev_value)
                                safe_write_cell(worksheet, f"{cols['growth']}{row}", round(growth, 2))
                                
                                # Process Repeat Customer Order Count
                                if 'weekly_old_orders' in df.columns:
                                    current_orders = df.iloc[1]['weekly_old_orders']
                                    prev_orders = df.iloc[0]['weekly_old_orders']
                                    orders_growth = ((float(current_orders) - float(prev_orders)) / float(prev_orders)) * 100 if prev_orders != 0 else 0
                                    
                                    row_orders = row_mappings['Customers']['Repeat Customer Order Count']
                                    safe_write_cell(worksheet, f"{cols['data']}{row_orders}", current_orders)
                                    safe_write_cell(worksheet, f"{cols['prev']}{row_orders}", prev_orders)
                                    safe_write_cell(worksheet, f"{cols['growth']}{row_orders}", round(orders_growth, 2))
                                    
                            except Exception as e:
                                logger.error(f"Error processing repeat_customers for {region}: {str(e)}")
                                logger.error(traceback.format_exc())

                    # Discount Orders
                    elif query_name == 'discount_orders':
                        if not df.empty:
                            try:
                                logger.info(f"Processing discount_orders for {region}")
                                logger.info(f"DataFrame columns: {df.columns.tolist()}")
                                logger.info(f"DataFrame first row: {df.iloc[0].to_dict() if len(df) > 0 else 'Empty'}")
                                
                                discount_mappings = {
                                    'No_Discount_Orders': 'No Discount Orders',
                                    'Rest_sponsored_Orders': 'Restaurant Sponsored Orders',
                                    'Smiles_Sponsored_Orders': 'Smiles Sponsored Orders',
                                    'Cofund_Orders': 'Co-fund orders',
                                    'Flat_Discount_Orders': 'Flat Discount'
                                }
                                
                                for df_col, param_name in discount_mappings.items():
                                    if df_col in df.columns:
                                        row = row_mappings['Discounts'][param_name]
                                        current_value = df.iloc[0][df_col]
                                        prev_value = df.iloc[1][df_col] if len(df) > 1 else 0
                                        
                                        safe_write_cell(worksheet, f"{cols['data']}{row}", current_value)
                                        safe_write_cell(worksheet, f"{cols['prev']}{row}", prev_value)
                                        
                                        if prev_value != 0:
                                            growth = ((float(current_value) - float(prev_value)) / float(prev_value)) * 100
                                            safe_write_cell(worksheet, f"{cols['growth']}{row}", round(growth, 2))
                                    else:
                                        logger.warning(f"Column {df_col} not found in discount_orders DataFrame")
                                        
                            except Exception as e:
                                logger.error(f"Error processing discount_orders for {region}: {str(e)}")
                                logger.error(f"DataFrame info: {df.info()}")
                                logger.error(traceback.format_exc())

                except Exception as e:
                    logger.error(f"Error processing query {query_name} for region {region}: {str(e)}")
                    logger.error(traceback.format_exc())
                    raise

        except Exception as e:
            logger.error(f"Error in _process_query_results for region {region}: {str(e)}")
            logger.error(traceback.format_exc())
            raise

    def parse_queries(self, content: str) -> Dict[str, str]:
        """Parse queries from the content string."""
        # Split on any sequence of dashes that's 10 or more characters long
        query_blocks = re.split(r'-{10,}', content)
        
        # Clean up each block and remove empty ones
        query_blocks = [block.strip() for block in query_blocks if block.strip()]
        
        queries = {}
        query_mapping = {
            'Query 1': 'orders_metrics',
            'Query 2': 'sales_metrics',
            'Query 3': 'commission_metrics',
            'Query 4': 'order_frequency',
            'Query 5': 'subscription_orders',
            'Query 6': 'new_customers',
            'Query 7': 'repeat_customers',
            'Query 8': 'repeat_orders',
            'Query 9': 'discount_orders'
        }
        
        # Map each query block to its corresponding key
        for block in query_blocks:
            for query_id, query_key in query_mapping.items():
                if query_id in block:
                    queries[query_key] = block

        # Verify all queries were found
        for query_key in query_mapping.values():
            if query_key not in queries or not queries[query_key].strip():
                raise ValueError(f"Missing or empty query: {query_key}")
        
        return queries

def main():
    try:
        logger.info("Starting Weekly Report Generation")
        
        # Check if template exists
        template_path = 'WoW_report_template.xlsx'
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Excel template not found: {template_path}")
        
        # Check template file permissions
        if not os.access(template_path, os.R_OK):
            raise PermissionError(f"No read permission for template file: {template_path}")
        
        # Database configuration from environment variables
        db_config = {
            'ssh_host': os.environ['SSH_HOST'],
            'ssh_username': os.environ['SSH_USERNAME'],
            'mysql_host': os.environ['MYSQL_HOST'],
            'mysql_user': os.environ['MYSQL_USER'],
            'mysql_password': os.environ['MYSQL_PASSWORD'],
            'mysql_database': os.environ['MYSQL_DATABASE']
        }
        
        # Initialize report generator
        generator = ReportGenerator(db_config)
        
        # Read and parse queries
        query_file_path = 'Queries for Weekly Report.txt'
        if not os.path.exists(query_file_path):
            raise FileNotFoundError(f"Query file not found: {query_file_path}")
        
        with open(query_file_path, 'r', encoding='utf-8') as file:
            content = file.read()
            
        # Parse queries and generate report
        queries = generator.parse_queries(content)
        
        # Process regions and generate report
        results = {}
        for region in generator.regions:
            try:
                logger.info(f"Processing region: {region}")
                region_result, region_data = generator.process_region(region, queries)
                results[region] = region_data
                logger.info(f"Successfully processed region: {region}")
            except Exception as e:
                logger.error(f"Failed to process region {region}: {str(e)}")
                raise
        
        # Generate the report
        output_dir = 'output'
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, 'Weekly Report.xlsx')
        
        # Update Excel template with results
        generator.update_excel_template(template_path, results, output_file)
        
        # Verify file exists before sending
        if not os.path.exists(output_file):
            raise FileNotFoundError(f"Generated report not found at: {output_file}")
            
        logger.info(f"Report generated successfully at: {output_file}")
        
        # Send email with the report
        send_email(output_file)
        
        logger.info("Weekly report generation and email sending completed successfully")
        
    except Exception as e:
        logger.error("Error during weekly report generation")
        logger.error(str(e))
        logger.error(traceback.format_exc())
        raise
    finally:
        logger.info("Script execution completed")

if __name__ == "__main__":
    main()